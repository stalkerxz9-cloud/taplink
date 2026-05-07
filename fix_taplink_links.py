"""
fix_taplink_links.py
────────────────────
Автоматически исправляет неверные ссылки во всех аккаунтах Taplink.

Трансформация URL:
  ❌  https://auto.ae/en/SLUG/?category=CATEGORY
  ✅  https://auto.ae/SLUG/CATEGORY/

Требования:
  pip install playwright openpyxl
  playwright install chromium
"""

import re
import asyncio
import logging
from pathlib import Path

import openpyxl
from playwright.async_api import async_playwright, Page, TimeoutError as PWTimeout

# ──────────────────────────────────────────────
# НАСТРОЙКИ
# ──────────────────────────────────────────────
XLSX_PATH    = "тaplink.xlsx"          # путь к вашей таблице
LOG_FILE     = "fix_taplink.log"
HEADLESS     = True                    # False — видеть браузер
CONCURRENCY  = 3                       # параллельных вкладок (не больше 5)
TIMEOUT      = 15_000                  # мс на ожидание элементов

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)

# ──────────────────────────────────────────────
# ТРАНСФОРМАЦИЯ URL
# ──────────────────────────────────────────────
BAD_PATTERN = re.compile(
    r"https://auto\.ae/en/([^/?]+)/?\?category=([^&\s]+)",
    re.IGNORECASE,
)

def fix_url(url: str) -> str | None:
    """
    Возвращает исправленный URL или None если URL уже верный / не подходит.
    """
    m = BAD_PATTERN.search(url.strip())
    if not m:
        return None
    slug, category = m.group(1), m.group(2)
    return f"https://auto.ae/{slug}/{category}/"

# ──────────────────────────────────────────────
# ЧТЕНИЕ ТАБЛИЦЫ
# ──────────────────────────────────────────────
def load_accounts() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c else "" for c in rows[0]]
    accounts = []
    for row in rows[1:]:
        rec = dict(zip(header, row))
        taplink = str(rec.get("Ссылка Taplink", "") or "").strip()
        email   = str(rec.get("Email", "") or "").strip()
        pwd     = str(rec.get("Пароль", "") or "").strip()
        name    = str(rec.get("Название шоурума", "") or "").strip()
        if email and pwd and taplink:
            accounts.append({
                "name":    name,
                "taplink": taplink,
                "email":   email,
                "password": pwd,
            })
    log.info(f"Загружено аккаунтов: {len(accounts)}")
    return accounts

# ──────────────────────────────────────────────
# ЛОГИН
# ──────────────────────────────────────────────
async def login(page: Page, email: str, password: str) -> bool:
    try:
        await page.goto("https://taplink.cc/p/login", timeout=TIMEOUT)
        await page.fill('input[name="email"]',    email,    timeout=TIMEOUT)
        await page.fill('input[name="password"]', password, timeout=TIMEOUT)
        await page.click('button[type="submit"]', timeout=TIMEOUT)
        # ждём редиректа на дашборд
        await page.wait_for_url("**/p/**", timeout=TIMEOUT)
        return True
    except PWTimeout:
        log.warning(f"  Логин не удался (timeout): {email}")
        return False
    except Exception as e:
        log.warning(f"  Логин ошибка: {e}")
        return False

# ──────────────────────────────────────────────
# ИСПРАВЛЕНИЕ ССЫЛОК НА СТРАНИЦЕ РЕДАКТОРА
# ──────────────────────────────────────────────
async def fix_account(page: Page, acc: dict) -> dict:
    name   = acc["name"]
    email  = acc["email"]
    result = {"name": name, "email": email, "fixed": 0, "status": "ok", "note": ""}

    ok = await login(page, email, acc["password"])
    if not ok:
        result["status"] = "error"
        result["note"]   = "login failed"
        return result

    try:
        # Открываем редактор
        await page.goto("https://taplink.cc/p/blocks/", timeout=TIMEOUT)
        await page.wait_for_load_state("networkidle", timeout=TIMEOUT)

        # Собираем все блоки с кнопками/ссылками
        # Taplink рендерит блоки — ищем все редактируемые поля URL
        link_inputs = await page.query_selector_all('input[data-type="url"], input[placeholder*="http"], input[name*="url"], input[name*="link"]')

        fixed_count = 0
        for inp in link_inputs:
            val = await inp.input_value()
            new_val = fix_url(val)
            if new_val:
                log.info(f"  [{name}] Исправляю: {val}  →  {new_val}")
                await inp.triple_click()
                await inp.type(new_val)
                fixed_count += 1

        if fixed_count:
            # Ищем кнопку «Сохранить»
            save_btn = page.locator('button:has-text("Сохранить"), button:has-text("Save")')
            if await save_btn.count() > 0:
                await save_btn.first.click(timeout=TIMEOUT)
                await page.wait_for_load_state("networkidle", timeout=TIMEOUT)
                log.info(f"  [{name}] Сохранено, исправлено ссылок: {fixed_count}")
            else:
                result["note"] = "save button not found"
        else:
            log.info(f"  [{name}] Неверных ссылок не найдено")

        result["fixed"] = fixed_count

    except PWTimeout:
        result["status"] = "error"
        result["note"]   = "timeout in editor"
        log.error(f"  [{name}] Timeout в редакторе")
    except Exception as e:
        result["status"] = "error"
        result["note"]   = str(e)
        log.error(f"  [{name}] Ошибка: {e}")

    # Разлогиниться перед следующим аккаунтом
    try:
        await page.goto("https://taplink.cc/p/logout", timeout=TIMEOUT)
    except Exception:
        pass

    return result

# ──────────────────────────────────────────────
# ПАРАЛЛЕЛЬНЫЙ ЗАПУСК
# ──────────────────────────────────────────────
async def worker(semaphore, browser, acc, all_results):
    async with semaphore:
        page = await browser.new_page()
        try:
            r = await fix_account(page, acc)
            all_results.append(r)
        finally:
            await page.close()

async def main():
    accounts = load_accounts()
    results  = []
    sem      = asyncio.Semaphore(CONCURRENCY)

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=HEADLESS)
        tasks = [worker(sem, browser, acc, results) for acc in accounts]
        await asyncio.gather(*tasks)
        await browser.close()

    # ──────────────────────────────────────────
    # ИТОГОВЫЙ ОТЧЁТ
    # ──────────────────────────────────────────
    import openpyxl as xl
    wb = xl.Workbook()
    ws = wb.active
    ws.title = "Результат"
    ws.append(["Название", "Email", "Исправлено ссылок", "Статус", "Примечание"])
    ok_count  = sum(1 for r in results if r["status"] == "ok")
    err_count = sum(1 for r in results if r["status"] == "error")
    for r in results:
        ws.append([r["name"], r["email"], r["fixed"], r["status"], r["note"]])
    wb.save("taplink_fix_report.xlsx")

    log.info("=" * 50)
    log.info(f"Готово. Успешно: {ok_count}, Ошибок: {err_count}")
    log.info("Отчёт сохранён: taplink_fix_report.xlsx")

if __name__ == "__main__":
    asyncio.run(main())
